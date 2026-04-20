"""
Unit tests for Excel export functionality.

Tests the HighwaySegmentationExcelExporter class including data parsing,
worksheet creation, formatting, and error handling.
"""

import pytest
import sys
import os
import json
import pandas as pd
import tempfile
import shutil
from pathlib import Path
from unittest.mock import Mock, patch, mock_open

# Add src to path for imports - portable approach
current_file_dir = os.path.dirname(__file__)  # tests/unit
tests_dir = os.path.dirname(current_file_dir)  # tests  
project_root = os.path.dirname(tests_dir)  # highway-segmentation-ga
src_path = os.path.join(project_root, 'src')

# Add to path if not already present
if src_path not in sys.path:
    sys.path.insert(0, src_path)

try:
    from excel_export import HighwaySegmentationExcelExporter, export_json_to_excel
    import openpyxl
except ImportError as e:
    pytest.skip(f"Required modules not available: {e}", allow_module_level=True)


@pytest.fixture
def sample_json_data():
    """Sample JSON data for testing Excel export."""
    return {
        "analysis_metadata": {
            "timestamp": "2026-04-10T12:00:00",
            "analysis_method": "single",
            "analysis_status": "completed",
            "software_version": {
                "application": "Highway Segmentation",
                "version": "1.95.2"
            },
            "input_file_info": {
                "data_file_path": "test_data.csv",
                "data_file_name": "test_data.csv",
                "total_data_rows": 100,
                "total_routes_available": 1,
                "column_info": {
                    "x_column": "BDFO", 
                    "y_column": "D60",
                    "route_column": "RDB"
                }
            }
        },
        "input_parameters": {
            "optimization_method_config": {
                "method_key": "single",
                "display_name": "Single-Objective GA"
            },
            "method_parameters": {
                "min_length": 2.0,
                "max_length": 5.0,
                "population_size": 50,
                "num_generations": 100
            }
        },
        "route_results": [{
            "route_info": {"route_id": "TEST_ROUTE"},
            "input_data_analysis": {
                "data_summary": {
                    "total_data_points": 100,
                    "data_range": {"x_min": 0.0, "x_max": 10.0}
                },
                "gap_analysis": {
                    "total_gaps": 0,
                    "gap_segments": []
                },
                "mandatory_segments": {
                    "mandatory_breakpoints": [0.0, 10.0],
                    "analyzable_segments": [{
                        "start": 0.0,
                        "end": 10.0,
                        "length": 10.0,
                        "type": "data"
                    }]
                }
            },
            "processing_results": {
                "best_solution": {
                    "objective_values": [3, 3.33],
                    "segmentation": {
                        "breakpoints": [0.0, 3.33, 6.67, 10.0],
                        "segment_count": 3,
                        "segments": [
                            {"start": 0.0, "end": 3.33, "length": 3.33},
                            {"start": 3.33, "end": 6.67, "length": 3.34},
                            {"start": 6.67, "end": 10.0, "length": 3.33}
                        ]
                    }
                }
            }
        }]
    }

@pytest.fixture
def sample_multi_route_json():
    """Sample multi-route JSON data for testing."""
    return {
        "analysis_metadata": {
            "timestamp": "2026-04-10T12:00:00",
            "analysis_method": "multi",
            "analysis_status": "completed",
            "software_version": {"application": "Highway Segmentation", "version": "1.95.2"}
        },
        "route_results": [
            {"route_info": {"route_id": "ROUTE_A"}, 
             "processing_results": {"pareto_points": [
                 {"point_id": 0, "objective_values": [2, 5.0], 
                  "segmentation": {"breakpoints": [0.0, 5.0, 10.0]}}
             ]}},
            {"route_info": {"route_id": "ROUTE_B"}, 
             "processing_results": {"pareto_points": [
                 {"point_id": 0, "objective_values": [3, 3.33], 
                  "segmentation": {"breakpoints": [0.0, 3.33, 6.67, 10.0]}}
             ]}}
        ]
    }

@pytest.fixture
def sample_csv_data():
    """Sample CSV data for testing."""
    return pd.DataFrame({
        'BDFO': [0.0, 1.0, 2.0, 3.0, 4.0],
        'D60': [5.0, 4.5, 6.0, 3.5, 5.5],
        'RDB': ['TEST_ROUTE'] * 5
    })


class TestHighwaySegmentationExcelExporter:
    """Test suite for Excel export functionality."""
    
    @pytest.mark.unit
    def test_init_basic(self, sample_json_data):
        """Test basic initialization of exporter."""
        exporter = HighwaySegmentationExcelExporter(sample_json_data)
        assert exporter.json_data == sample_json_data
        assert exporter.original_csv_path is None
        assert hasattr(exporter, 'workbook')

    @pytest.mark.unit
    def test_init_with_csv_path(self, sample_json_data):
        """Test initialization with CSV path."""
        csv_path = "test.csv"
        exporter = HighwaySegmentationExcelExporter(sample_json_data, csv_path)
        assert exporter.original_csv_path == csv_path
        assert hasattr(exporter, 'workbook')

    @pytest.mark.unit
    def test_init_with_nonexistent_csv(self, sample_json_data):
        """Test initialization with non-existent CSV file."""
        exporter = HighwaySegmentationExcelExporter(sample_json_data, "nonexistent.csv")
        assert exporter.original_csv_path == "nonexistent.csv"
        assert hasattr(exporter, 'workbook')

    @pytest.mark.unit
    def test_worksheet_creation(self, sample_json_data):
        """Test worksheet creation with headers."""
        exporter = HighwaySegmentationExcelExporter(sample_json_data)
        
        headers = ["Header1", "Header2", "Header3"]
        ws = exporter._create_worksheet_with_headers("Test Sheet", headers)
        
        assert ws.title == "Test Sheet"
        assert ws['A1'].value == "Header1"
        assert ws['B1'].value == "Header2" 
        assert ws['C1'].value == "Header3"
        
        # Headers should be bold
        assert ws['A1'].font.bold is True

    @pytest.mark.unit
    def test_analysis_summary_tab_creation(self, sample_json_data):
        """Test creation of analysis summary worksheet."""
        exporter = HighwaySegmentationExcelExporter(sample_json_data)
        exporter._create_analysis_summary_tab()
        
        # Verify worksheet was created
        assert 'Analysis Summary' in exporter.workbook.sheetnames
        ws = exporter.workbook['Analysis Summary']
        
        # Check headers
        assert ws['A1'].value == 'Category'
        assert ws['B1'].value == 'Item'
        
        # Should have some content rows
        assert ws.max_row > 1

    @pytest.mark.unit
    def test_export_to_excel_success(self, sample_json_data, tmp_path):
        """Test complete Excel export process."""
        output_file = tmp_path / "test_export.xlsx"
        exporter = HighwaySegmentationExcelExporter(sample_json_data)
        
        success, _ = exporter.export_to_excel(str(output_file))
        
        assert success is True
        assert output_file.exists()
        assert output_file.stat().st_size > 0
        
        # Verify worksheets were created
        wb = openpyxl.load_workbook(output_file)
        expected_sheets = [
            'Analysis Summary', 'Input Parameters', 'Route Summary',
            'Mandatory Breakpoints & Gaps', 'All Solutions', 'All Segmentation Output',
            'Analyzable Segments', 'Original Data', 'Statistics & Performance'
        ]
        
        for sheet_name in expected_sheets:
            assert sheet_name in wb.sheetnames
        
        wb.close()

    @pytest.mark.unit
    def test_export_to_excel_invalid_path(self, sample_json_data):
        """Test export with invalid output path."""
        exporter = HighwaySegmentationExcelExporter(sample_json_data)
        
        # Try to write to invalid directory
        invalid_path = "/invalid/path/that/does/not/exist/output.xlsx"
        success, _ = exporter.export_to_excel(invalid_path)
        
        assert success is False

    @pytest.mark.unit
    def test_export_json_to_excel_function(self, sample_json_data, tmp_path):
        """Test the standalone export function."""
        # Write JSON to temporary file
        json_file = tmp_path / "test_input.json"
        with open(json_file, 'w') as f:
            json.dump(sample_json_data, f)
        
        output_file = tmp_path / "test_output.xlsx"
        
        success = export_json_to_excel(str(json_file), str(output_file))
        
        assert success is True
        assert output_file.exists()

    @pytest.mark.unit
    def test_export_json_to_excel_invalid_json(self, tmp_path):
        """Test export function with invalid JSON file."""
        # Create invalid JSON file
        json_file = tmp_path / "invalid.json"
        with open(json_file, 'w') as f:
            f.write("{ invalid json content")
        
        output_file = tmp_path / "output.xlsx"
        
        success = export_json_to_excel(str(json_file), str(output_file))
        assert success is False

    @pytest.mark.unit  
    def test_large_dataset_handling(self, tmp_path):
        """Test export with large dataset to verify performance."""
        # Create large JSON data
        large_json = {
            "analysis_metadata": {"timestamp": "2026-04-10T12:00:00"},
            "route_results": []
        }
        
        # Add many routes with many Pareto points
        for route_num in range(5):
            route_data = {
                "route_info": {"route_id": f"ROUTE_{route_num}"},
                "processing_results": {"pareto_points": []}
            }
            
            # Add many Pareto points per route
            for point_num in range(20):
                point = {
                    "point_id": point_num,
                    "objective_values": [point_num, point_num * 1.5],
                    "segmentation": {
                        "breakpoints": list(range(point_num + 2)),
                        "segments": [{"start": i, "end": i+1, "length": 1} 
                                   for i in range(point_num + 1)]
                    }
                }
                route_data["processing_results"]["pareto_points"].append(point)
            
            large_json["route_results"].append(route_data)
        
        output_file = tmp_path / "large_export.xlsx"
        exporter = HighwaySegmentationExcelExporter(large_json)
        
        # Should handle large dataset without errors
        success, _ = exporter.export_to_excel(str(output_file))
        assert success is True
        assert output_file.exists()

    @pytest.mark.unit
    def test_special_characters_handling(self, tmp_path):
        """Test handling of special characters in data."""
        special_json = {
            "analysis_metadata": {
                "timestamp": "2026-04-10T12:00:00",
                "analysis_method": "test_with_special_chars_éñ中文"
            },
            "route_results": [{
                "route_info": {"route_id": "Route with spaces & symbols!@#"},
                "processing_results": {"pareto_points": []}
            }]
        }
        
        output_file = tmp_path / "special_chars.xlsx"
        exporter = HighwaySegmentationExcelExporter(special_json)
        
        success, _ = exporter.export_to_excel(str(output_file))
        assert success is True
        
        # Verify special characters are preserved
        wb = openpyxl.load_workbook(output_file)
        ws = wb['Route Summary']
        assert "Route with spaces & symbols!@#" in str(ws['A2'].value)
        wb.close()


@pytest.mark.unit
class TestExcelExportEdgeCases:
    """Test edge cases and error conditions."""
    
    def test_empty_json_data(self, tmp_path):
        """Test export with minimal/empty JSON data."""
        empty_json = {"analysis_metadata": {}, "route_results": []}
        
        output_file = tmp_path / "empty.xlsx"
        exporter = HighwaySegmentationExcelExporter(empty_json)
        
        # Should not crash, but create minimal Excel file
        success, _ = exporter.export_to_excel(str(output_file))
        assert success is True

    def test_missing_keys_json(self, tmp_path):
        """Test export with JSON missing expected keys."""
        incomplete_json = {"some_other_key": "value"}
        
        output_file = tmp_path / "incomplete.xlsx"
        exporter = HighwaySegmentationExcelExporter(incomplete_json)
        
        # Should handle gracefully
        success, _ = exporter.export_to_excel(str(output_file))
        assert success is True

    def test_csv_load_permission_error(self, sample_json_data):
        """Test handling of CSV file permission errors."""
        # Just test that initialization doesn't crash even with invalid path
        exporter = HighwaySegmentationExcelExporter(sample_json_data, "protected.csv")
        assert exporter.original_csv_path == "protected.csv"
        # Should still be able to export without CSV data
            

# Performance and integration markers
pytestmark = [pytest.mark.unit, pytest.mark.excel_export]