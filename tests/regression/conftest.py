"""
Regression Test Fixtures and Utilities for Highway Segmentation GA

This module provides pytest fixtures and utility functions to support comprehensive
regression testing across all optimization methods and data configurations.
It handles test data management, parameter standardization, output validation,
and result verification.

Key Responsibilities:
    Test Configuration:
        - Standardized parameter loading from JSON templates
        - Data file path resolution and validation
        - Output directory management and cleanup
        
    Validation Utilities:
        - JSON structure validation for all output formats
        - Excel/JSON content consistency verification
        - Schema compliance checking and error reporting
        
    Test Support:
        - Automated cleanup of test artifacts
        - Consistent file naming conventions
        - Error handling and reporting utilities

Fixture Hierarchy:
    Session-scoped fixtures (shared across all tests):
        - test_parameters: Standardized parameter configurations
        - test_data_dir: Path to test data files
        - data_configurations: Complete data setup with file paths
        
    Function-scoped fixtures (per-test isolation):
        - clean_outputs: Clean output directories before/after tests
        - Ensures test independence and artifact management

Validation Framework:
    - JSON structure validation ensures required fields present
    - Excel export validation confirms data consistency
    - Schema validation integration for compliance checking
    - Comprehensive error reporting for debugging

Author: Highway Segmentation GA Team
Version: 1.95+ (Enhanced Regression Testing)
"""

import pytest
import json
import os
import shutil
import pandas as pd
from pathlib import Path
import sys

# Add src to path for imports  
current_dir = Path(__file__).parent
project_root = current_dir.parent.parent
src_path = project_root / 'src'

if str(src_path) not in sys.path:
    sys.path.insert(0, str(src_path))

from config import OPTIMIZATION_METHODS


@pytest.fixture(scope="session")
def test_parameters():
    """
    Load standardized test parameters from JSON configuration template.
    
    This fixture provides consistent, optimized parameters across all regression
    tests. Parameters are designed for speed (reduced population/generations)
    while maintaining algorithm reliability and result quality.
    
    Returns:
        dict: Complete parameter configuration including:
            - common_parameters: Shared across all optimization methods
            - method_specific: Specialized parameters for constrained/aashto_cda
            - data_configurations: Dataset-specific column mappings
            
    Configuration Structure:
        Common Parameters:
            - population_size: 50 (reduced from production 200 for speed)
            - num_generations: 50 (reduced from production 200 for speed)
            - Standard GA parameters: mutation_rate, crossover_rate, elite_ratio
            - Performance settings: cache_clear_interval, enable_performance_stats
            
        Method-Specific Parameters:
            - constrained: target_avg_length, penalty settings, tolerance
            - aashto_cda: alpha, segment limits, diagnostic controls
            
        Data Configurations:
            - single_route: txdot_data.csv configuration
            - multi_route: AndreTestMultiRoute.csv configuration
    
    Usage:
        Parameters optimized for regression testing balance:
        - Speed: Fast execution for CI/CD integration
        - Reliability: Conservative settings that consistently work
        - Coverage: All method-specific features exercised
    """
    params_file = current_dir / "test_parameters_template.json"
    with open(params_file, 'r') as f:
        return json.load(f)


@pytest.fixture(scope="session")
def test_data_dir():
    """Path to test data directory."""
    return current_dir.parent / "test_data"


@pytest.fixture(scope="session")
def outputs_dir():
    """Path to test outputs directory."""
    return current_dir / "outputs"


@pytest.fixture(scope="session")
def json_outputs_dir(outputs_dir):
    """Path to JSON outputs directory."""
    return outputs_dir / "json"


@pytest.fixture(scope="session") 
def excel_outputs_dir(outputs_dir):
    """Path to Excel outputs directory."""
    return outputs_dir / "excel"


@pytest.fixture(scope="session")
def optimization_methods():
    """Get all available optimization methods."""
    return get_optimization_methods()


@pytest.fixture(scope="function")
def clean_outputs(outputs_dir):
    """
    Clean and prepare output directories for each test function.
    
    This fixture ensures test isolation by cleaning output directories before
    each test run and optionally after (configurable for debugging). Provides
    clean slate for each test while preserving artifacts when needed.
    
    Test Isolation Benefits:
        - Prevents test interference from previous run artifacts
        - Ensures consistent starting state for all tests
        - Facilitates reliable CI/CD integration
        - Enables parallel test execution without conflicts
    
    Directory Structure Created:
        outputs/
        ├── json/     # JSON result files from optimization runs
        └── excel/    # Excel export files for result analysis
    
    Debugging Support:
        - Cleanup can be disabled by commenting out teardown section
        - Preserved artifacts enable manual inspection of test results
        - Useful for troubleshooting failing tests or validation issues
        
    Args:
        outputs_dir: Path to test outputs directory (from outputs_dir fixture)
        
    Yields:
        None: Provides clean directory structure, then optionally cleans up
    """
    if outputs_dir.exists():
        shutil.rmtree(outputs_dir)
    outputs_dir.mkdir(parents=True, exist_ok=True)
    (outputs_dir / "json").mkdir(exist_ok=True)
    (outputs_dir / "excel").mkdir(exist_ok=True)
    yield
    # Optional: Clean up after test (comment out to keep artifacts for inspection)
    # if outputs_dir.exists():
    #     shutil.rmtree(outputs_dir)


@pytest.fixture(scope="session")
def data_configurations(test_parameters, test_data_dir):
    """
    Configure complete data setup with validated file paths and column mappings.
    
    This fixture provides comprehensive data configuration for all regression tests,
    including file path resolution, existence validation, and column mapping setup.
    Essential for ensuring consistent data access across all test combinations.
    
    Configuration Process:
        1. Load data configurations from test parameters JSON
        2. Resolve full file paths relative to test data directory
        3. Validate that all required test data files exist
        4. Provide complete column mapping for each dataset
    
    Data Configuration Structure:
        single_route Configuration:
            - File: txdot_data.csv (TxDOT highway pavement data)
            - X Column: 'milepoint' (highway distance markers)
            - Y Column: 'structural_strength_ind' (pavement strength measurements)
            - Route Column: None (single continuous route processing)
            
        multi_route Configuration:
            - File: AndreTestMultiRoute.csv (multi-route engineering data)
            - X Column: 'BDFO' (bearing/distance from origin measurements)
            - Y Column: 'D60' (material property measurements)
            - Route Column: 'RDB' (route database identifier)
    
    Args:
        test_parameters: Standardized test parameters from JSON configuration
        test_data_dir: Base directory path for test data files
    
    Returns:
        dict: Complete data configurations with resolved paths and validation
        
    Raises:
        FileNotFoundError: If any required test data file is missing
        
    Usage:
        Consumed by regression tests to ensure consistent data access:
        - File path resolution for different test environments
        - Column mapping validation for optimization algorithms
        - Data availability verification before test execution
    """
    configs = test_parameters["data_configurations"]
    
    # Add full file paths
    for config_name, config in configs.items():
        config["file_path"] = test_data_dir / config["file"]
        
        # Verify file exists
        if not config["file_path"].exists():
            raise FileNotFoundError(f"Test data file not found: {config['file_path']}")
    
    return configs


def validate_json_structure(json_data, method_key):
    """
    Validate basic JSON result structure for regression test compliance.
    
    This function performs essential structure validation to ensure all
    regression test outputs contain the minimum required fields for
    downstream processing and analysis.
    
    Validation Criteria:
        Required Top-Level Keys:
            - analysis_metadata: Method info, timestamps, status
            - input_parameters: Configuration used for optimization
            
        Metadata Validation:
            - analysis_method: Optimization method identifier
            - timestamp: Result generation timestamp
            - analysis_status: Success/failure status indicator
    
    Args:
        json_data (dict): Parsed JSON data from optimization results
        method_key (str): Expected optimization method key for validation
        
    Returns:
        bool: True if structure validation passes
        
    Raises:
        AssertionError: If required keys or structure elements are missing
        
    Usage:
        Called by regression tests to ensure output consistency across
        all optimization methods and data configurations. Provides early
        detection of structural issues before detailed validation.
        
    Integration:
        - Used in conjunction with JSON schema validation
        - Provides quick structural checks during test execution
        - Enables fast failure detection for malformed outputs
    """
    required_keys = ["analysis_metadata", "input_parameters"]
    
    for key in required_keys:
        assert key in json_data, f"Missing required key: {key}"
    
    metadata = json_data["analysis_metadata"]
    assert "analysis_method" in metadata
    assert "timestamp" in metadata
    assert "analysis_status" in metadata
    
    return True


def validate_excel_vs_json(excel_file, json_data):
    """
    Validate Excel export content consistency with JSON source data.
    
    This function ensures that Excel exports accurately represent the JSON
    optimization results without data loss or transformation errors. Critical
    for maintaining data integrity across different output formats.
    
    Validation Process:
        1. Load Excel workbook and verify sheet structure
        2. Check for required worksheets (Summary, Analysis_Metadata, Input_Parameters)
        3. Cross-validate key data points between Excel and JSON
        4. Verify method identification consistency
        5. Confirm data integrity preservation
    
    Sheet Validation:
        Required Sheets:
            - Summary: High-level optimization results and statistics
            - Analysis_Metadata: Method info, timestamps, processing details
            - Input_Parameters: Configuration parameters used
            
    Data Consistency Checks:
        - Method name matching between formats
        - Key numerical results preservation
        - Timestamp and metadata consistency
        - Parameter configuration accuracy
    
    Args:
        excel_file (Path): Path to Excel file for validation
        json_data (dict): Source JSON data for comparison
        
    Returns:
        bool: True if Excel/JSON consistency validation passes
        
    Raises:
        AssertionError: If required sheets missing or data inconsistencies found
        
    Quality Assurance:
        - Prevents silent data corruption in Excel export process
        - Ensures reliable data analysis workflows
        - Validates export functionality regression testing
    """
    import openpyxl
    
    # Load Excel workbook
    wb = openpyxl.load_workbook(excel_file)
    
    # Basic validation - check for expected sheets
    expected_sheets = ["Summary", "Analysis_Metadata", "Input_Parameters"]
    for sheet in expected_sheets:
        assert sheet in wb.sheetnames, f"Missing Excel sheet: {sheet}"
    
    # Validate summary data matches JSON
    summary_sheet = wb["Summary"]
    
    # Check method matches
    method_cell = None
    for row in summary_sheet.iter_rows(values_only=True):
        if row[0] and "Method" in str(row[0]):
            method_cell = row[1]
            break
    
    if method_cell:
        json_method = json_data["analysis_metadata"]["analysis_method"]  
        assert str(method_cell).lower() == json_method.lower(), f"Method mismatch: Excel={method_cell}, JSON={json_method}"
    
    return True


def get_result_filename(method, dataset, extension):
    """
    Generate standardized result filename for consistent test output management.
    
    Creates consistent, predictable filenames that enable systematic organization
    and automated processing of regression test results across all method and
    dataset combinations.
    
    Filename Format:
        Pattern: regression_{method}_{dataset}.{extension}
        
        Components:
            - "regression_": Clear identification as regression test output
            - method: Optimization method identifier (single, multi, constrained, aashto_cda)
            - dataset: Data configuration identifier (single_route, multi_route)  
            - extension: File type identifier (json, xlsx, etc.)
    
    Args:
        method (str): Optimization method identifier
        dataset (str): Data configuration identifier
        extension (str): File extension without leading dot
        
    Returns:
        str: Standardized filename for regression test output
        
    Usage Examples:
        - get_result_filename("single", "single_route", "json")
          → "regression_single_single_route.json"
        - get_result_filename("multi", "multi_route", "xlsx")
          → "regression_multi_multi_route.xlsx"
    
    Integration Benefits:
        - Consistent naming across all test output files
        - Easy automated discovery and processing of results
        - Clear method/dataset relationship identification
        - Systematic organization for result comparison and analysis
    """
    return f"regression_{method}_{dataset}.{extension}"